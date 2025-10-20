# -*- coding: utf-8 -*-
"""
Created on Fri Sep 26 15:33:07 2025

@author: meissnerto
"""

from pathlib import Path
from fnmatch import fnmatch
from datetime import datetime
import re
import os
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from carlquant_frames.specimen_model import (
    Specimen, SliceResult, RegionStats, LesionDepth, Surface,
    SpecimenConfig, RegionConfig, AirConfig
)
from PIL import Image, ImageDraw
import numpy as np


def convert_to_json_serializable(obj):
    """
    Recursively convert numpy types to native Python types for JSON serialization.
    
    Args:
        obj: Object to convert (can be dict, list, numpy type, or native type)
    
    Returns:
        JSON-serializable version of the object
    """
    if isinstance(obj, dict):
        return {key: convert_to_json_serializable(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [convert_to_json_serializable(item) for item in obj]
    elif isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, (np.bool_, bool)):
        return bool(obj)
    else:
        return obj



IMAGE_EXTENSIONS = ['*.jpg', '*.png', '*.tif', '*.tiff']

def natural_key(path):
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split(r'(\d+)', path.name)]

class DataLoader:
    @staticmethod
    def find_image_stacks(root_folder: Path) -> dict[str, Specimen]:
        specimen_data = {}
        for subdir in root_folder.rglob("*"):
            if subdir.is_dir():
                # Skip 'annotations' folders - they contain processed images with overlays
                if subdir.name.lower() == "annotations":
                    continue
                
                # Skip if any parent folder is named 'annotations'
                if any(parent.name.lower() == "annotations" for parent in subdir.parents):
                    continue
                
                image_files = sorted([
                    f for f in subdir.iterdir()
                    if f.is_file() and any(fnmatch(f.name.lower(), ext) for ext in IMAGE_EXTENSIONS)
                ], key=natural_key)

                data_folders = [f for f in subdir.iterdir() if f.is_dir() and f.name.startswith("Data_")]

                if image_files:
                    specimen_id = subdir.name
                    specimen = Specimen(
                        specimen_id=subdir.name,
                        source=subdir,
                        images=image_files,
                        slices=len(image_files),
                        regions="",
                        status="New",
                        date=subdir.stat().st_mtime,
                        previous_runs=data_folders
                    )
                    
                    # Don't load configuration here - it will be loaded after metadata is set
                    # This ensures we load from the correct Data_{operator}_{measurement} folder
                    
                    specimen_data[specimen_id] = specimen
        return specimen_data

    @staticmethod
    def load_specimen_config(specimen: Specimen) -> SpecimenConfig:
        """Load specimen configuration from JSON file if it exists.
        
        Also loads computed annotations (surface, lesion_depth, extraction_regions) if available.
        Prioritizes loading from Data_{operator}_{measurement} folder if specimen has metadata.
        """
        try:
            # If specimen has operator/measurement metadata, look for specific folder first
            if hasattr(specimen, 'operator') and hasattr(specimen, 'measurement'):
                target_folder = specimen.source / f"Data_{specimen.operator}_{specimen.measurement}"
                if target_folder.exists():
                    config_file = target_folder / f"{specimen.specimen_id}_config.json"
                    if config_file.exists():
                        with open(config_file, 'r') as f:
                            config_data = json.load(f)
                        return DataLoader._parse_config_data(specimen, config_data)
            
            # Fallback: Look for config file in any Data_ folder (legacy behavior)
            for data_folder in specimen.previous_runs:
                config_file = data_folder / f"{specimen.specimen_id}_config.json"
                if config_file.exists():
                    with open(config_file, 'r') as f:
                        config_data = json.load(f)
                    return DataLoader._parse_config_data(specimen, config_data)
            
            return None
        except Exception:
            return None
    
    @staticmethod
    def _parse_config_data(specimen: Specimen, config_data: dict) -> SpecimenConfig:
        """Parse config data from JSON into SpecimenConfig object.
        
        Args:
            specimen: Specimen object
            config_data: Dictionary loaded from JSON
        
        Returns:
            SpecimenConfig object
        """
        # Parse the JSON data into our data structures
        config = SpecimenConfig(specimen_id=specimen.specimen_id)
        
        # Load regions
        if 'regions' in config_data:
            for slice_idx_str, region_data in config_data['regions'].items():
                slice_idx = int(slice_idx_str)
                # Support both old (2-point) and new (4-point) format
                if 'specimen_start' in region_data:
                    # New 4-point format
                    config.regions[slice_idx] = RegionConfig(
                        slice_index=slice_idx,
                        specimen_start=tuple(region_data['specimen_start']),
                        lesion_start=tuple(region_data['lesion_start']),
                        lesion_end=tuple(region_data['lesion_end']),
                        tooth_end=tuple(region_data['tooth_end']),
                        is_keyframe=region_data.get('is_keyframe', False)  # Default False for backward compatibility
                    )
                else:
                    # Old 2-point format - convert to 4-point
                    # Assume: specimen_start = start_point, lesion_start = start_point,
                    #         lesion_end = end_point, tooth_end = end_point
                    start_pt = tuple(region_data['start_point'])
                    end_pt = tuple(region_data['end_point'])
                    config.regions[slice_idx] = RegionConfig(
                        slice_index=slice_idx,
                        specimen_start=start_pt,
                        lesion_start=start_pt,
                        lesion_end=end_pt,
                        tooth_end=end_pt,
                        is_keyframe=False  # Old format, mark as not keyframe
                    )
        
        # Load air configurations
        if 'air' in config_data:
            for slice_idx_str, air_data in config_data['air'].items():
                slice_idx = int(slice_idx_str)
                point2 = tuple(air_data['point2']) if air_data.get('point2') else None
                config.air[slice_idx] = AirConfig(
                    slice_index=slice_idx,
                    point1=tuple(air_data['point1']),
                    is_keyframe=air_data.get('is_keyframe', False),  # Default False for backward compatibility
                    point2=point2
                )
        
        # Load computed annotations if available
        if 'annotations' in config_data:
            DataLoader._load_annotations_into_results(specimen, config_data['annotations'])
        
        return config
    
    @staticmethod
    def _load_annotations_into_results(specimen: Specimen, annotations_data: dict):
        """Load computed annotations from JSON into specimen.results.
        
        Args:
            specimen: Specimen object to populate
            annotations_data: Dictionary of annotations keyed by slice index
        """
        for slice_idx_str, slice_annotations in annotations_data.items():
            slice_idx = int(slice_idx_str)
            
            # Initialize result containers
            surface = None
            lesion_depth = None
            region_stats = []
            
            # Load surface detection results
            if 'surface' in slice_annotations:
                surface_data = slice_annotations['surface']
                raw_points = [tuple(pt) for pt in surface_data.get('raw_points', [])]
                fitted_curves = {}
                for curve_name, points in surface_data.get('fitted_curves', {}).items():
                    fitted_curves[curve_name] = [tuple(pt) for pt in points]
                
                surface = Surface(
                    raw_points=raw_points,
                    fitted_curves=fitted_curves,
                    is_cavitated=surface_data.get('is_cavitated', False),
                    cavitation_depth=surface_data.get('cavitation_depth', 0.0)
                )
            
            # Load lesion depth results
            if 'lesion_depth' in slice_annotations:
                ld_data = slice_annotations['lesion_depth']
                depth_points = [tuple(pt) for pt in ld_data.get('depth_points', [])]
                smoothed_points = None
                if 'smoothed_depth_points' in ld_data:
                    smoothed_points = [tuple(pt) for pt in ld_data['smoothed_depth_points']]
                
                # Load lesion_detection_data if available (for debug visualization)
                lesion_detection_data = None
                if 'lesion_detection_data' in ld_data:
                    # Convert string keys back to integers
                    lesion_detection_data = {
                        int(x): data for x, data in ld_data['lesion_detection_data'].items()
                    }
                
                lesion_depth = LesionDepth(
                    depth_points=depth_points,
                    mean_depth=ld_data.get('mean_depth', 0.0),
                    median_depth=ld_data.get('median_depth', 0.0),
                    sd=ld_data.get('sd', 0.0),
                    se=ld_data.get('se', 0.0),
                    smoothed_depth_points=smoothed_points,
                    lesion_detection_data=lesion_detection_data
                )
            
            # Load extraction region bounds and statistics
            if 'extraction_regions' in slice_annotations:
                for region_data in slice_annotations['extraction_regions']:
                    # Parse bounds
                    bounds = None
                    if 'bounds' in region_data and region_data['bounds']:
                        bounds_data = region_data['bounds']
                        if isinstance(bounds_data[0], list):
                            # Rotated corners
                            bounds = tuple(tuple(pt) for pt in bounds_data)
                        else:
                            # Simple bbox
                            bounds = tuple(bounds_data)
                    
                    region_stats.append(RegionStats(
                        region_type=region_data.get('region_type', 'sound'),
                        pixel_values=[],  # Not saved in annotations (too large)
                        mean=region_data.get('mean', 0.0),
                        median=region_data.get('median', 0.0),
                        sd=region_data.get('sd', 0.0),
                        se=region_data.get('se', 0.0),
                        region_index=region_data.get('region_index', 0),
                        bounds=bounds if bounds else (0, 0, 0, 0),
                        rotation_angle=region_data.get('rotation_angle', 0.0)
                    ))
            
            # Store the slice result if we have any data
            if surface or lesion_depth or region_stats:
                specimen.results[slice_idx] = SliceResult(
                    slice_index=slice_idx,
                    region_stats=region_stats,
                    surface=surface if surface else Surface([], {}),
                    lesion_depth=lesion_depth if lesion_depth else LesionDepth([], 0, 0, 0, 0)
                )

    @staticmethod
    def load_results(specimen: Specimen, region_config: dict):
        """Load results from Excel file, respecting operator/measurement metadata.
        
        Only loads from Data_{operator}_{measurement} folder matching the specimen's metadata.
        This ensures results from different analysis sessions don't get mixed up.
        """
        if not specimen.previous_runs:
            return

        try:
            # Use metadata-specific folder instead of just picking the latest
            # This matches the behavior of load_specimen_config()
            target_folder = None
            
            if hasattr(specimen, 'operator') and hasattr(specimen, 'measurement'):
                # Look for specific Data_{operator}_{measurement} folder
                target_folder = specimen.source / f"Data_{specimen.operator}_{specimen.measurement}"
                if not target_folder.exists():
                    return  # No matching folder for this operator/measurement
            else:
                # Fallback: use latest folder (legacy behavior, should rarely happen)
                target_folder = max(specimen.previous_runs, key=lambda p: p.stat().st_mtime)
            
            result_file = target_folder / f"{specimen.specimen_id}_results.xlsx"
            if not result_file.exists():
                return

            wb = load_workbook(result_file)
            if "Summary" not in wb.sheetnames:
                return

            ws = wb["Summary"]
            sound_count = region_config.get("sound", 3)
            lesion_count = region_config.get("lesion", 3)
            expected_columns = 1 + sound_count + lesion_count + 1  # slice + regions + depth

            specimen.results.clear()
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if not row or len(row) < expected_columns:
                    continue

                try:
                    # Excel now uses 1-based slice numbers, convert back to 0-based for internal storage
                    slice_index_from_excel = row[0] if isinstance(row[0], int) else (i + 1)
                    slice_index = slice_index_from_excel - 1  # Convert to 0-based
                    sound_medians = row[1 : 1 + sound_count]
                    lesion_medians = row[1 + sound_count : 1 + sound_count + lesion_count]
                    lesion_depth_mean = row[1 + sound_count + lesion_count]

                    region_stats = [
                        RegionStats("sound", [], mean=0, median=float(m), sd=0, se=0)
                        for m in sound_medians
                    ] + [
                        RegionStats("lesion", [], mean=0, median=float(m), sd=0, se=0)
                        for m in lesion_medians
                    ]

                    specimen.results[slice_index] = SliceResult(
                        slice_index=slice_index,
                        region_stats=region_stats,
                        surface=Surface([], {}),
                        lesion_depth=LesionDepth([], mean_depth=lesion_depth_mean, median_depth=0, sd=0, se=0)
                    )

                except Exception:
                    continue

        except Exception:
            pass


class DataSaver:
    @staticmethod
    def store_slice_result(specimen: Specimen, slice_index: int,
                           region_stats: list[RegionStats],
                           surface: Surface,
                           lesion_depth: LesionDepth):
        specimen.results[slice_index] = SliceResult(
            slice_index=slice_index,
            region_stats=region_stats,
            surface=surface,
            lesion_depth=lesion_depth
        )

    @staticmethod
    def save_results(specimen: Specimen):
        wb = Workbook()

        # === Sheet 1: Summary ===
        ws_summary = wb.active
        ws_summary.title = "Summary"

        num_sound = sum(1 for r in specimen.results[0].region_stats if r.region_type == "sound")
        num_lesion = sum(1 for r in specimen.results[0].region_stats if r.region_type == "lesion")

        headers = ["SLICE"]
        headers += [f"SOUND_{i+1}_MEDIAN" for i in range(num_sound)]
        headers += [f"LESION_{i+1}_MEDIAN" for i in range(num_lesion)]
        headers += ["LESION_DEPTH_MEAN", "IS_CAVITATED"]
        ws_summary.append(headers)

        # Sort by slice_index (ascending) and use 1-based numbering for humans
        for slice_index, result in sorted(specimen.results.items(), key=lambda x: x[0]):
            row = [slice_index + 1]
            row += [r.median for r in result.region_stats if r.region_type == "sound"]
            row += [r.median for r in result.region_stats if r.region_type == "lesion"]
            row += [result.lesion_depth.mean_depth if result.lesion_depth else 0]
            
            # IS_CAVITATED: Use "TRUE"/"FALSE" strings for consistency, blank if no data
            if result.surface and hasattr(result.surface, 'is_cavitated'):
                cavitated_value = "TRUE" if result.surface.is_cavitated else "FALSE"
            else:
                cavitated_value = ""  # Leave blank if no surface data available
            row += [cavitated_value]
            
            ws_summary.append(row)

        # === Sheet 2: Region Pixels ===
        ws_pixels = wb.create_sheet("Region Pixels")
        
        # Create headers: SLICE, PIXEL_INDEX, SOUND_1..SOUND_N, LESION_1..LESION_N
        pixel_headers = ["SLICE", "PIXEL_INDEX"]
        pixel_headers += [f"SOUND_{i+1}" for i in range(num_sound)]
        pixel_headers += [f"LESION_{i+1}" for i in range(num_lesion)]
        ws_pixels.append(pixel_headers)
        
        # Transpose data: one row per pixel instead of one row per region
        # Sort by slice_index (ascending) and use 1-based numbering
        for slice_index, result in sorted(specimen.results.items(), key=lambda x: x[0]):
            # Get all regions for this slice
            sound_regions = [r for r in result.region_stats if r.region_type == "sound"]
            lesion_regions = [r for r in result.region_stats if r.region_type == "lesion"]
            
            # Determine max pixel count across all regions
            max_pixels = max(
                max((len(r.pixel_values) for r in sound_regions), default=0),
                max((len(r.pixel_values) for r in lesion_regions), default=0)
            )
            
            # Write one row per pixel
            for pixel_idx in range(max_pixels):
                row = [slice_index + 1, pixel_idx + 1]
                
                # Add sound region values for this pixel
                for sound_region in sound_regions:
                    if pixel_idx < len(sound_region.pixel_values):
                        row.append(sound_region.pixel_values[pixel_idx])
                    else:
                        row.append(None)  # Empty cell if region has fewer pixels
                
                # Add lesion region values for this pixel
                for lesion_region in lesion_regions:
                    if pixel_idx < len(lesion_region.pixel_values):
                        row.append(lesion_region.pixel_values[pixel_idx])
                    else:
                        row.append(None)  # Empty cell if region has fewer pixels
                
                ws_pixels.append(row)
        
        # MEMORY CLEANUP: Clear pixel_values after writing to Excel
        # Pixel values are large (625 per region × 12 regions × N slices)
        # They're preserved in Excel and not needed in memory anymore
        for result in specimen.results.values():
            for stats in result.region_stats:
                stats.pixel_values.clear()

        # === Sheet 3: Surface & Depth ===
        ws_surface = wb.create_sheet("Surface & Depth")
        ws_surface.append(["SLICE", "TYPE", "X", "Y"])

        # Sort by slice_index (ascending) and use 1-based numbering
        for slice_index, result in sorted(specimen.results.items(), key=lambda x: x[0]):
            for x, y in result.surface.raw_points:
                ws_surface.append([slice_index + 1, "raw_surface", x, y])
            for curve_name, points in result.surface.fitted_curves.items():
                for x, y in points:
                    ws_surface.append([slice_index + 1, f"curve_{curve_name}", x, y])
            if result.lesion_depth and result.lesion_depth.depth_points:
                for x, y in result.lesion_depth.depth_points:
                    ws_surface.append([slice_index + 1, "lesion_depth", x, y])

        # === Save to disk ===
        operator = getattr(specimen, "operator", "OP")
        measurement = getattr(specimen, "measurement", 1)
        save_folder = specimen.source / f"Data_{operator}_{measurement}"
        save_folder.mkdir(exist_ok=True)

        target_file = save_folder / f"{specimen.specimen_id}_results.xlsx"
        wb.save(target_file)
        
        # Save annotations to config JSON (includes surface, lesion_depth, extraction_regions)
        DataSaver.save_specimen_config(specimen, include_annotations=True)

    @staticmethod
    def save_specimen_config(specimen: Specimen, include_annotations: bool = False):
        """Save specimen configuration (REGIONS, AIR, and optionally computed annotations) to JSON file.
        
        Args:
            specimen: Specimen object to save
            include_annotations: If True, save computed annotations (surface, lesion_depth, extraction_regions)
        """
        if not specimen.config:
            return
        
        # Create save folder if it doesn't exist
        operator = getattr(specimen, "operator", "OP")
        measurement = getattr(specimen, "measurement", 1)
        save_folder = specimen.source / f"Data_{operator}_{measurement}"
        save_folder.mkdir(exist_ok=True)
        
        # Prepare data for JSON serialization
        config_data = {
            "specimen_id": specimen.config.specimen_id,
            "regions": {},
            "air": {}
        }
        
        # Convert regions to JSON-serializable format
        for slice_idx, region_config in specimen.config.regions.items():
            config_data["regions"][str(slice_idx)] = {
                "slice_index": region_config.slice_index,
                "specimen_start": list(region_config.specimen_start),
                "lesion_start": list(region_config.lesion_start),
                "lesion_end": list(region_config.lesion_end),
                "tooth_end": list(region_config.tooth_end),
                "is_keyframe": region_config.is_keyframe
            }
        
        # Convert air configurations to JSON-serializable format
        for slice_idx, air_config in specimen.config.air.items():
            air_data = {
                "slice_index": air_config.slice_index,
                "point1": list(air_config.point1),
                "is_keyframe": air_config.is_keyframe
            }
            if air_config.point2:
                air_data["point2"] = list(air_config.point2)
            config_data["air"][str(slice_idx)] = air_data
        
        # Save computed annotations if requested and available
        if include_annotations and hasattr(specimen, 'results') and specimen.results:
            config_data["annotations"] = {}
            
            for slice_idx, result in specimen.results.items():
                slice_annotations = {}
                
                # Save surface detection results
                if result.surface:
                    surface_data = {
                        "raw_points": [[int(x), int(y)] for x, y in result.surface.raw_points],
                        "fitted_curves": {}
                    }
                    for curve_name, points in result.surface.fitted_curves.items():
                        surface_data["fitted_curves"][curve_name] = [[int(x), int(y)] for x, y in points]
                    
                    # Save cavitation detection if available
                    if hasattr(result.surface, 'is_cavitated'):
                        surface_data["is_cavitated"] = bool(result.surface.is_cavitated)
                        surface_data["cavitation_depth"] = float(result.surface.cavitation_depth)
                    
                    slice_annotations["surface"] = surface_data
                
                # Save lesion depth results
                if result.lesion_depth:
                    lesion_depth_data = {
                        "depth_points": [[int(x), int(y)] for x, y in result.lesion_depth.depth_points],
                        "mean_depth": float(result.lesion_depth.mean_depth),
                        "median_depth": float(result.lesion_depth.median_depth),
                        "sd": float(result.lesion_depth.sd),
                        "se": float(result.lesion_depth.se)
                    }
                    
                    # Save smoothed depth points if available
                    if hasattr(result.lesion_depth, 'smoothed_depth_points') and result.lesion_depth.smoothed_depth_points:
                        lesion_depth_data["smoothed_depth_points"] = [[int(x), int(y)] for x, y in result.lesion_depth.smoothed_depth_points]
                    
                    # Save lesion_detection_data for debug visualization of component methods
                    if hasattr(result.lesion_depth, 'lesion_detection_data') and result.lesion_depth.lesion_detection_data:
                        # Convert lesion_detection_data to JSON-serializable format
                        # Only save essential metadata, not full intensity profiles (too large)
                        detection_data_serializable = {}
                        for x, data in result.lesion_depth.lesion_detection_data.items():
                            detection_data_serializable[str(x)] = {
                                'surface_y': int(data.get('surface_y', 0)),
                                'actual_depth': float(data.get('actual_depth', np.nan)) if not np.isnan(data.get('actual_depth', np.nan)) else None,
                                'knee_depth': float(data.get('knee_depth', np.nan)) if not np.isnan(data.get('knee_depth', np.nan)) else None,
                                'detection_metadata': convert_to_json_serializable(data.get('detection_metadata', {}))
                            }
                        lesion_depth_data["lesion_detection_data"] = detection_data_serializable
                    
                    slice_annotations["lesion_depth"] = lesion_depth_data
                
                # Save extraction region bounds and statistics
                if result.region_stats:
                    regions_data = []
                    for stats in result.region_stats:
                        region_data = {
                            "region_type": str(stats.region_type),
                            "region_index": int(stats.region_index),
                            "mean": float(stats.mean),
                            "median": float(stats.median),
                            "sd": float(stats.sd),
                            "se": float(stats.se),
                            "rotation_angle": float(stats.rotation_angle)
                        }
                        
                        # Save bounds (can be 4 corner points or simple bbox)
                        if stats.bounds:
                            if isinstance(stats.bounds[0], tuple):
                                # Rotated corners
                                region_data["bounds"] = [[int(x), int(y)] for x, y in stats.bounds]
                            else:
                                # Simple bbox
                                region_data["bounds"] = [int(v) for v in stats.bounds]
                        
                        regions_data.append(region_data)
                    
                    slice_annotations["extraction_regions"] = regions_data
                
                if slice_annotations:
                    config_data["annotations"][str(slice_idx)] = slice_annotations
        
        # Save to JSON file
        config_file = save_folder / f"{specimen.specimen_id}_config.json"
        with open(config_file, 'w') as f:
            json.dump(config_data, f, indent=2)

    @staticmethod
    def update_specimen_region(specimen: Specimen, slice_index: int, 
                              specimen_start: tuple, lesion_start: tuple,
                              lesion_end: tuple, tooth_end: tuple,
                              context=None, auto_save=True, is_keyframe=False):
        """Update region configuration for a specific slice (4 points).
        
        Args:
            specimen: Specimen to update
            slice_index: Slice index
            specimen_start: Specimen start point
            lesion_start: Lesion start point
            lesion_end: Lesion end point
            tooth_end: Tooth end point
            context: Application context (for metadata)
            auto_save: If True, automatically save config to JSON after update
            is_keyframe: If True, marks this as a user-defined keyframe (not interpolated)
        """
        if not specimen.config:
            specimen.config = SpecimenConfig(specimen_id=specimen.specimen_id)
        
        specimen.config.regions[slice_index] = RegionConfig(
            slice_index=slice_index,
            specimen_start=specimen_start,
            lesion_start=lesion_start,
            lesion_end=lesion_end,
            tooth_end=tooth_end,
            is_keyframe=is_keyframe
        )
        
        # Store metadata in specimen if available from context
        if context:
            metadata = getattr(context, "analysis_metadata", {})
            operator = metadata.get("operator", "OP")
            measurement = metadata.get("measurement", 1)
            specimen.operator = operator
            specimen.measurement = measurement
        
        # Auto-save configuration if requested
        if auto_save:
            DataSaver.save_specimen_config(specimen)

    @staticmethod
    def update_specimen_air(specimen: Specimen, slice_index: int, point1: tuple, point2: tuple = None, context=None, auto_save=True, is_keyframe=False):
        """Update AIR configuration for a specific slice.
        
        Args:
            specimen: Specimen to update
            slice_index: Slice index
            point1: First point
            point2: Second point
            context: Application context (for metadata)
            is_keyframe: If True, marks this as a user-defined keyframe (not interpolated)
            auto_save: If True, automatically save config to JSON after update
        """
        if not specimen.config:
            specimen.config = SpecimenConfig(specimen_id=specimen.specimen_id)
        
        specimen.config.air[slice_index] = AirConfig(
            slice_index=slice_index,
            point1=point1,
            point2=point2,
            is_keyframe=is_keyframe
        )
        
        # Store metadata in specimen if available from context
        if context:
            metadata = getattr(context, "analysis_metadata", {})
            operator = metadata.get("operator", "OP")
            measurement = metadata.get("measurement", 1)
            specimen.operator = operator
            specimen.measurement = measurement
        
        # Auto-save configuration if requested
        if auto_save:
            DataSaver.save_specimen_config(specimen)

    @staticmethod
    def save_annotated_images(specimen: Specimen):
        """
        Save images with annotations overlaid for visualization.
        
        Creates an 'annotations' folder inside Data_{operator}_{measurement}
        and saves each slice with surface, lesion depth, regions, and boundaries drawn.
        Images are saved with their original filenames (e.g., 'tooth_001.png') in the
        annotations folder, making it easy to identify which original image was processed.
        
        Args:
            specimen: Specimen object with results and config
        """
        if not specimen.results:
            return
        
        # Create annotations folder
        operator = getattr(specimen, "operator", "OP")
        measurement = getattr(specimen, "measurement", 1)
        save_folder = specimen.source / f"Data_{operator}_{measurement}" / "annotations"
        save_folder.mkdir(parents=True, exist_ok=True)
        
        # Process each slice
        for slice_idx, result in specimen.results.items():
            try:
                # Load original image
                if slice_idx >= len(specimen.images):
                    continue
                
                image_path = specimen.images[slice_idx]
                img = Image.open(image_path).convert('RGB')  # Convert to RGB for colored annotations
                draw = ImageDraw.Draw(img)
                
                # Get configurations for this slice
                region_config = specimen.config.regions.get(slice_idx) if specimen.config else None
                air_config = specimen.config.air.get(slice_idx) if specimen.config else None
                
                # Draw AIR region (cyan rectangle)
                if air_config and air_config.point2:
                    x1, y1 = air_config.point1
                    x2, y2 = air_config.point2
                    draw.rectangle([x1, y1, x2, y2], outline='cyan', width=2)
                
                # Draw region boundaries (vertical lines)
                if region_config:
                    # Specimen start (green)
                    x = region_config.specimen_start[0]
                    draw.line([(x, 0), (x, img.height)], fill='#00FF66', width=2)
                    
                    # Lesion start (yellow)
                    x = region_config.lesion_start[0]
                    draw.line([(x, 0), (x, img.height)], fill='yellow', width=2)
                    
                    # Lesion end (yellow)
                    x = region_config.lesion_end[0]
                    draw.line([(x, 0), (x, img.height)], fill='yellow', width=2)
                    
                    # Tooth end (green)
                    x = region_config.tooth_end[0]
                    draw.line([(x, 0), (x, img.height)], fill='#00FF66', width=2)
                
                # Draw surface curves
                if result.surface and result.surface.fitted_curves:
                    # Interpolated surface curve (cyan) - if cavitated
                    if "interpolated_surface" in result.surface.fitted_curves:
                        points = result.surface.fitted_curves["interpolated_surface"]
                        if len(points) > 1:
                            for i in range(len(points) - 1):
                                draw.line([points[i], points[i+1]], fill='cyan', width=1)
                    
                    # Actual surface curve (orange)
                    if "actual_surface" in result.surface.fitted_curves:
                        points = result.surface.fitted_curves["actual_surface"]
                        if len(points) > 1:
                            for i in range(len(points) - 1):
                                draw.line([points[i], points[i+1]], fill='orange', width=2)
                
                # Draw lesion depth
                if result.lesion_depth:
                    # Use smoothed points if available, otherwise raw points
                    points = None
                    if hasattr(result.lesion_depth, 'smoothed_depth_points') and result.lesion_depth.smoothed_depth_points:
                        points = result.lesion_depth.smoothed_depth_points
                    elif result.lesion_depth.depth_points:
                        points = result.lesion_depth.depth_points
                    
                    if points and len(points) > 1:
                        # Draw line
                        for i in range(len(points) - 1):
                            draw.line([points[i], points[i+1]], fill='red', width=2)
                
                # Draw extraction regions
                if result.region_stats:
                    from PIL import ImageFont
                    try:
                        font = ImageFont.truetype("arial.ttf", 12)
                    except:
                        font = ImageFont.load_default()
                    
                    for stats in result.region_stats:
                        if not stats.bounds or len(stats.bounds) == 0:
                            continue
                        
                        color = '#00FF66' if stats.region_type == "sound" else 'red'
                        
                        # Check if rotated corners or simple bbox
                        if len(stats.bounds) == 4 and isinstance(stats.bounds[0], tuple):
                            # Rotated rectangle - draw polygon
                            corners = list(stats.bounds) + [stats.bounds[0]]  # Close the polygon
                            draw.line(corners, fill=color, width=2)
                            
                            # Calculate center for label
                            center_x = sum(x for x, y in stats.bounds) / 4
                            center_y = sum(y for x, y in stats.bounds) / 4
                        else:
                            # Simple rectangle
                            left_x, top_y, right_x, bottom_y = stats.bounds
                            draw.rectangle([left_x, top_y, right_x, bottom_y], outline=color, width=2)
                            
                            center_x = (left_x + right_x) / 2
                            center_y = (top_y + bottom_y) / 2
                        
                        # Draw region number
                        text = str(stats.region_index)
                        bbox = draw.textbbox((center_x, center_y), text, font=font)
                        text_width = bbox[2] - bbox[0]
                        text_height = bbox[3] - bbox[1]
                        draw.text((center_x - text_width/2, center_y - text_height/2), text, fill=color, font=font)
                
                # Save annotated image using original filename (without _annotated suffix since folder is named "annotations")
                original_filename = specimen.images[slice_idx].stem  # Get filename without extension
                output_filename = f"{original_filename}.png"
                output_path = save_folder / output_filename
                img.save(output_path, 'PNG')
                
            except Exception as e:
                # Log error but continue with other slices
                print(f"Error saving annotated image for slice {slice_idx}: {e}")
                continue


